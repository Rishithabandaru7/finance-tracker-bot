import os
import io
import threading
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from collections import defaultdict
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler,
    MessageHandler, filters, ContextTypes
)
from dotenv import load_dotenv
from database import (
    init_db, register_user, add_transaction,
    get_monthly_summary, set_budget, check_budget,
    get_recent_transactions, delete_transaction,
    delete_all_transactions, get_all_transactions,
    get_all_users, add_savings_goal, update_savings_goal,
    get_savings_goals, add_bill_reminder, get_bill_reminders,
    delete_bill_reminder, add_custom_category,
    get_custom_categories, add_split_expense, get_split_expenses
)
from groq_parser import parse_message
from scheduler import start_scheduler, build_summary_message

load_dotenv()

# ── Health Check Server ───────────────────────────────────

class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"OK")
    def log_message(self, format, *args):
        pass

def run_health_server():
    server = HTTPServer(("0.0.0.0", 8080), HealthHandler)
    server.serve_forever()

# ── Helper ────────────────────────────────────────────────

def get_user(update: Update):
    user = update.effective_user
    register_user(user.id, user.username, user.first_name)
    return user.id, user.first_name

# ── Start ─────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    msg = f"""👋 *Welcome {first_name} to Finance Tracker Bot!*

I understand natural language — just type like you're texting a friend! 🤖

━━━━━━━━━━━━━━━━━━━━━
💸 *TRACK EXPENSES*
━━━━━━━━━━━━━━━━━━━━━
Just type naturally:
- `spent 500 on lunch`
- `paid 1200 for uber`
- `bought medicines 350`
- `rent 10000`

Multiple at once:
- `food 500, uber 1200, rent 10000`

━━━━━━━━━━━━━━━━━━━━━
💰 *TRACK INCOME*
━━━━━━━━━━━━━━━━━━━━━
- `received 50000 salary`
- `got 5000 freelance payment`
- `business income 20000`

━━━━━━━━━━━━━━━━━━━━━
📊 *VIEW REPORTS*
━━━━━━━━━━━━━━━━━━━━━
- /summary — monthly breakdown
- /recent — last 5 transactions
- /chart — spending pie chart
- /export — download Excel report

━━━━━━━━━━━━━━━━━━━━━
🎯 *BUDGET MANAGEMENT*
━━━━━━━━━━━━━━━━━━━━━
- `/budget food 5000` — set ₹5000 food budget
- Auto alert at 80% and 100% usage!

━━━━━━━━━━━━━━━━━━━━━
🎯 *SAVINGS GOALS*
━━━━━━━━━━━━━━━━━━━━━
- `/goal add iPhone 80000` — create goal
- `/goal save iPhone 5000` — add savings
- `/goals` — view all goals with progress bar

━━━━━━━━━━━━━━━━━━━━━
🔔 *BILL REMINDERS*
━━━━━━━━━━━━━━━━━━━━━
- `/bill add Electricity 1500 5` — due on 5th
- `/bills` — view all bills
- Auto reminder 3 days, 1 day & on due date!

━━━━━━━━━━━━━━━━━━━━━
👥 *SPLIT EXPENSES*
━━━━━━━━━━━━━━━━━━━━━
- `/split 3000 dinner John,Mary`
- Calculates each person's share instantly!

━━━━━━━━━━━━━━━━━━━━━
🏷️ *CUSTOM CATEGORIES*
━━━━━━━━━━━━━━━━━━━━━
- `/addcategory gym expense` — add category
- `/categories` — view all categories

━━━━━━━━━━━━━━━━━━━━━
🗑️ *DELETE TRANSACTIONS*
━━━━━━━━━━━━━━━━━━━━━
- `/recent` — see last 5 with IDs
- `/delete 5` — delete transaction #5
- `/delete all` — fresh start

━━━━━━━━━━━━━━━━━━━━━
📅 *AUTO WEEKLY SUMMARY*
━━━━━━━━━━━━━━━━━━━━━
Every Sunday 8PM you'll automatically
receive your weekly spending summary! 📬

━━━━━━━━━━━━━━━━━━━━━
🏦 *DEFAULT CATEGORIES*
━━━━━━━━━━━━━━━━━━━━━
Expenses: food, transport, shopping,
entertainment, health, bills, rent, other

Income: salary, freelance, business,
investment, other

━━━━━━━━━━━━━━━━━━━━━
🔒 Your data is *completely private*
Nobody else can see your transactions!
━━━━━━━━━━━━━━━━━━━━━

*Start now — type your first expense!* 💪
Example: `spent 500 on lunch`"""
    await update.message.reply_text(msg, parse_mode="Markdown")

# ── Summary ───────────────────────────────────────────────

async def summary_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    rows = get_monthly_summary(user_id)
    month = datetime.now().strftime("%B %Y")
    msg = build_summary_message(rows, f"{first_name}'s Summary — {month}")
    await update.message.reply_text(msg, parse_mode="Markdown")

# ── Budget ────────────────────────────────────────────────

async def budget_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    args = context.args
    if len(args) != 2:
        await update.message.reply_text("Usage: /budget <category> <amount>\nExample: /budget food 5000")
        return
    category = args[0].lower()
    try:
        limit = float(args[1])
        set_budget(user_id, category, limit)
        await update.message.reply_text(
            f"✅ Budget set!\n🏷️ Category: {category.title()}\n💰 Limit: ₹{limit:,.0f}/month"
        )
    except ValueError:
        await update.message.reply_text("❌ Amount must be a number.")

# ── Recent ────────────────────────────────────────────────

async def recent_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    rows = get_recent_transactions(user_id)
    if not rows:
        await update.message.reply_text("No transactions found!")
        return
    msg = "🧾 *Your Last 5 Transactions:*\n\n"
    for row in rows:
        id_, type_, amount, category, desc, date = row
        emoji = "💸" if type_ == "expense" else "💰"
        msg += f"{emoji} *#{id_}* — ₹{amount:,.0f} | {category.title()}\n"
        msg += f"   📝 {desc}\n"
        msg += f"   📅 {date}\n"
        msg += f"   `/delete {id_}`\n\n"
    await update.message.reply_text(msg, parse_mode="Markdown")

# ── Delete ────────────────────────────────────────────────

async def delete_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    args = context.args
    if len(args) == 0:
        await update.message.reply_text(
            "Usage:\n• `/delete <id>` — delete one\n• `/delete all` — delete everything",
            parse_mode="Markdown"
        )
        return
    if args[0].lower() == "all":
        count = delete_all_transactions(user_id)
        await update.message.reply_text(f"✅ All {count} transactions deleted! 🧹")
        return
    try:
        transaction_id = int(args[0])
        success = delete_transaction(user_id, transaction_id)
        if success:
            await update.message.reply_text(f"✅ Transaction #{transaction_id} deleted!")
        else:
            await update.message.reply_text(f"❌ Transaction #{transaction_id} not found!")
    except ValueError:
        await update.message.reply_text("❌ Invalid input. Use /delete 5 or /delete all")

# ── Export Excel ──────────────────────────────────────────

async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    await update.message.reply_text("⏳ Generating Excel report...")

    rows = get_all_transactions(user_id)
    if not rows:
        await update.message.reply_text("❌ No transactions to export!")
        return

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "All Transactions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F7A4D")

    headers = ["#", "Date", "Type", "Category", "Description", "Amount (₹)"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_income = 0
    total_expense = 0

    for row_idx, row in enumerate(rows, 2):
        id_, type_, amount, category, desc, date = row
        if type_ == "income":
            fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
            total_income += amount
        else:
            fill = PatternFill(fill_type="solid", fgColor="FFEBEE")
            total_expense += amount
        data = [row_idx - 1, str(date), type_.title(), category.title(), desc, amount]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.fill = fill

    for col, width in zip("ABCDEF", [5, 12, 10, 14, 30, 14]):
        ws1.column_dimensions[col].width = width

    ws2 = wb.create_sheet("Summary")
    net = total_income - total_expense
    summary_data = [
        ["Finance Summary", ""],
        ["", ""],
        ["Total Income", total_income],
        ["Total Expenses", total_expense],
        ["Net Savings", net],
        ["Total Transactions", len(rows)],
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=value)
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15

    ws3 = wb.create_sheet("Category Breakdown")
    for col, header in enumerate(["Category", "Type", "Total (₹)"], 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    category_totals = defaultdict(lambda: {"expense": 0, "income": 0})
    for _, type_, amount, category, _, _ in rows:
        category_totals[category][type_] += amount

    row_idx = 2
    for category, totals in sorted(category_totals.items()):
        for type_, amount in totals.items():
            if amount > 0:
                ws3.cell(row=row_idx, column=1, value=category.title())
                ws3.cell(row=row_idx, column=2, value=type_.title())
                ws3.cell(row=row_idx, column=3, value=amount)
                row_idx += 1

    filename = f"/tmp/finance_{first_name}_{datetime.now().strftime('%Y%m')}.xlsx"
    wb.save(filename)

    with open(filename, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=f"finance_{first_name}.xlsx",
            caption=f"📊 *Finance Report*\n\n💰 Income: ₹{total_income:,.0f}\n💸 Expenses: ₹{total_expense:,.0f}\n✅ Net: ₹{net:,.0f}",
            parse_mode="Markdown"
        )
    os.remove(filename)

# ── Chart ─────────────────────────────────────────────────

async def chart_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    await update.message.reply_text("⏳ Generating chart...")

    rows = get_monthly_summary(user_id)
    expense_rows = [(cat, amt) for type_, cat, amt in rows if type_ == "expense"]

    if not expense_rows:
        await update.message.reply_text("❌ No expenses this month to chart!")
        return

    labels = [cat.title() for cat, _ in expense_rows]
    amounts = [amt for _, amt in expense_rows]
    colors = ["#FF6B6B", "#4ECDC4", "#45B7D1", "#96CEB4", "#FFEAA7", "#DDA0DD", "#98D8C8", "#F7DC6F"]

    fig, ax = plt.subplots(figsize=(8, 6))
    wedges, texts, autotexts = ax.pie(
        amounts,
        labels=labels,
        autopct="%1.1f%%",
        colors=colors[:len(labels)],
        startangle=90
    )
    ax.set_title(f"{first_name}'s Spending — {datetime.now().strftime('%B %Y')}", fontsize=14, fontweight="bold")

    buf = io.BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    buf.seek(0)
    plt.close()

    await update.message.reply_photo(
        photo=buf,
        caption=f"📈 *Spending Breakdown*\n\nTotal: ₹{sum(amounts):,.0f}"
    )

# ── Savings Goals ─────────────────────────────────────────

async def goal_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    args = context.args

    if len(args) == 0:
        await update.message.reply_text(
            "Usage:\n"
            "• `/goal add iPhone 80000` — create goal\n"
            "• `/goal save iPhone 5000` — add savings\n"
            "• `/goals` — view all goals",
            parse_mode="Markdown"
        )
        return

    if args[0].lower() == "add" and len(args) >= 3:
        goal_name = args[1]
        try:
            target = float(args[2])
            add_savings_goal(user_id, goal_name, target)
            await update.message.reply_text(
                f"🎯 *Goal Created!*\n\n"
                f"📌 {goal_name}\n"
                f"💰 Target: ₹{target:,.0f}\n\n"
                f"Save money using `/goal save {goal_name} <amount>`",
                parse_mode="Markdown"
            )
        except ValueError:
            await update.message.reply_text("❌ Amount must be a number.")

    elif args[0].lower() == "save" and len(args) >= 3:
        goal_name = args[1]
        try:
            amount = float(args[2])
            update_savings_goal(user_id, goal_name, amount)
            goals = get_savings_goals(user_id)
            for name, target, saved in goals:
                if name.lower() == goal_name.lower():
                    percent = (saved / target) * 100
                    remaining = target - saved
                    emoji = "🎉" if saved >= target else "💪"
                    await update.message.reply_text(
                        f"{emoji} *Savings Updated!*\n\n"
                        f"📌 {name}\n"
                        f"💰 Saved: ₹{saved:,.0f} / ₹{target:,.0f}\n"
                        f"📊 Progress: {percent:.1f}%\n"
                        f"🎯 Remaining: ₹{remaining:,.0f}",
                        parse_mode="Markdown"
                    )
                    return
        except ValueError:
            await update.message.reply_text("❌ Amount must be a number.")
    else:
        await update.message.reply_text("❌ Invalid. Use /goal add <name> <amount>")

async def goals_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    goals = get_savings_goals(user_id)

    if not goals:
        await update.message.reply_text("No savings goals yet!\n\nCreate one: `/goal add iPhone 80000`", parse_mode="Markdown")
        return

    msg = "🎯 *Your Savings Goals:*\n\n"
    for name, target, saved in goals:
        percent = (saved / target) * 100
        filled = int(percent / 10)
        bar = "█" * filled + "░" * (10 - filled)
        emoji = "🎉" if saved >= target else "💪"
        msg += f"{emoji} *{name}*\n"
        msg += f"   [{bar}] {percent:.1f}%\n"
        msg += f"   ₹{saved:,.0f} / ₹{target:,.0f}\n\n"

    await update.message.reply_text(msg, parse_mode="Markdown")

# ── Bill Reminders ────────────────────────────────────────

async def bill_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    args = context.args

    if len(args) == 0:
        await update.message.reply_text(
            "Usage:\n"
            "• `/bill add Electricity 1500 5` — due on 5th\n"
            "• `/bill delete <id>` — remove reminder\n"
            "• `/bills` — view all bills",
            parse_mode="Markdown"
        )
        return

    if args[0].lower() == "add" and len(args) >= 4:
        bill_name = args[1]
        try:
            amount = float(args[2])
            due_day = int(args[3])
            if not 1 <= due_day <= 31:
                await update.message.reply_text("❌ Due day must be between 1 and 31")
                return
            add_bill_reminder(user_id, bill_name, amount, due_day)
            await update.message.reply_text(
                f"🔔 *Bill Reminder Added!*\n\n"
                f"📋 {bill_name}\n"
                f"💰 ₹{amount:,.0f}\n"
                f"📅 Due on: {due_day}th of every month\n\n"
                f"You'll get reminders 3 days, 1 day before and on due date!",
                parse_mode="Markdown"
            )
        except ValueError:
            await update.message.reply_text("❌ Invalid amount or day.")

    elif args[0].lower() == "delete" and len(args) >= 2:
        try:
            bill_id = int(args[1])
            success = delete_bill_reminder(user_id, bill_id)
            if success:
                await update.message.reply_text(f"✅ Bill reminder #{bill_id} deleted!")
            else:
                await update.message.reply_text(f"❌ Bill #{bill_id} not found!")
        except ValueError:
            await update.message.reply_text("❌ Invalid ID.")
    else:
        await update.message.reply_text("❌ Invalid. Use /bill add <name> <amount> <due_day>")

async def bills_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    bills = get_bill_reminders(user_id)

    if not bills:
        await update.message.reply_text("No bill reminders!\n\nAdd one: `/bill add Electricity 1500 5`", parse_mode="Markdown")
        return

    today = datetime.now().day
    msg = "🔔 *Your Bill Reminders:*\n\n"
    for id_, bill_name, amount, due_day in bills:
        days_left = due_day - today
        if days_left < 0:
            status = "✅ Paid this month"
        elif days_left == 0:
            status = "🚨 Due TODAY!"
        elif days_left == 1:
            status = "⚠️ Due tomorrow!"
        else:
            status = f"📅 Due in {days_left} days"
        msg += f"📋 *{bill_name}* (#{id_})\n"
        msg += f"   💰 ₹{amount:,.0f} | Due: {due_day}th\n"
        msg += f"   {status}\n\n"

    await update.message.reply_text(msg, parse_mode="Markdown")

# ── Custom Categories ─────────────────────────────────────

async def addcategory_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    args = context.args

    if len(args) != 2:
        await update.message.reply_text(
            "Usage: `/addcategory <name> <expense/income>`\n"
            "Example: `/addcategory gym expense`",
            parse_mode="Markdown"
        )
        return

    category_name = args[0].lower()
    type_ = args[1].lower()

    if type_ not in ["expense", "income"]:
        await update.message.reply_text("❌ Type must be 'expense' or 'income'")
        return

    add_custom_category(user_id, category_name, type_)
    await update.message.reply_text(
        f"✅ *Custom Category Added!*\n\n"
        f"🏷️ {category_name.title()} ({type_})\n\n"
        f"Now you can use it in messages!",
        parse_mode="Markdown"
    )

async def categories_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    custom = get_custom_categories(user_id)

    msg = "🏷️ *Default Categories:*\n\n"
    msg += "💸 Expense: food, transport, shopping, entertainment, health, bills, rent, other\n\n"
    msg += "💰 Income: salary, freelance, business, investment, other\n\n"

    if custom:
        msg += "✨ *Your Custom Categories:*\n\n"
        for name, type_ in custom:
            emoji = "💸" if type_ == "expense" else "💰"
            msg += f"{emoji} {name.title()} ({type_})\n"

    await update.message.reply_text(msg, parse_mode="Markdown")

# ── Split Expenses ─────────────────────────────────────────

async def split_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    args = context.args

    if len(args) < 3:
        await update.message.reply_text(
            "Usage: `/split <amount> <description> <person1,person2>`\n"
            "Example: `/split 3000 dinner John,Mary`",
            parse_mode="Markdown"
        )
        return

    try:
        total_amount = float(args[0])
        description = args[1]
        people = args[2].split(",")
        total_people = len(people) + 1  # include yourself
        per_person = total_amount / total_people

        split_with = ", ".join(people)
        add_split_expense(user_id, description, total_amount, split_with, per_person)

        msg = f"👥 *Split Expense Created!*\n\n"
        msg += f"📋 {description}\n"
        msg += f"💰 Total: ₹{total_amount:,.0f}\n"
        msg += f"👤 Split between: You, {split_with}\n"
        msg += f"💵 Each person pays: ₹{per_person:,.0f}\n\n"
        msg += "People who owe you:\n"
        for person in people:
            msg += f"  • {person.strip()}: ₹{per_person:,.0f}\n"

        await update.message.reply_text(msg, parse_mode="Markdown")

    except ValueError:
        await update.message.reply_text("❌ Amount must be a number.\nExample: `/split 3000 dinner John,Mary`", parse_mode="Markdown")

# ── Message Handler ───────────────────────────────────────

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id, first_name = get_user(update)
    user_text = update.message.text
    

    parsed_list = parse_message(user_text)

    for parsed in parsed_list:
        action = parsed.get("action", "unknown")
        if action == "get_summary":
            await summary_command(update, context)
            return
        if action == "set_budget":
            amount = parsed.get("amount")
            category = parsed.get("category", "other")
            if amount and category:
                set_budget(user_id, category, amount)
                await update.message.reply_text(
                    f"✅ Budget set!\n🏷️ {category.title()}: ₹{amount:,.0f}/month",
                    parse_mode="Markdown"
                )
            else:
                await update.message.reply_text("❌ Try: 'set food budget to 5000'")
            return

    transactions = [p for p in parsed_list if p.get("action") == "add_transaction"]

    if not transactions:
        await update.message.reply_text(
            "🤔 Didn't catch that.\n\nTry:\n"
            "• 'spent 200 on coffee'\n"
            "• 'received 30000 salary'\n"
            "• 'food 500, uber 1200, rent 10000'\n"
            "• 'show summary'"
        )
        return

    reply = f"✅ *Recorded {len(transactions)} transaction(s)!*\n\n"

    for t in transactions:
        type_ = t.get("type")
        amount = t.get("amount")
        category = t.get("category", "other")
        desc = t.get("description", user_text)

        add_transaction(user_id, type_, amount, category, desc)

        emoji = "💸" if type_ == "expense" else "💰"
        reply += f"{emoji} ₹{amount:,.0f} | {category.title()} | {desc}\n"

        if type_ == "expense":
            spent, limit = check_budget(user_id, category)
            if limit:
                percent = (spent / limit) * 100
                if percent >= 100:
                    reply += f"   🚨 *Budget exceeded! {percent:.0f}% used*\n"
                elif percent >= 80:
                    reply += f"   ⚠️ *Budget warning! {percent:.0f}% used*\n"

    await update.message.reply_text(reply, parse_mode="Markdown")

# ── Main ──────────────────────────────────────────────────

def main():
    init_db()
    scheduler = start_scheduler()
    print("✅ Database initialized")
    print("✅ Scheduler started")
    print("🤖 Multi-user bot is running...")

    thread = threading.Thread(target=run_health_server)
    thread.daemon = True
    thread.start()
    print("✅ Health server started")

    app = ApplicationBuilder().token(os.getenv("TELEGRAM_BOT_TOKEN")).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("summary", summary_command))
    app.add_handler(CommandHandler("budget", budget_command))
    app.add_handler(CommandHandler("recent", recent_command))
    app.add_handler(CommandHandler("delete", delete_command))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("chart", chart_command))
    app.add_handler(CommandHandler("goal", goal_command))
    app.add_handler(CommandHandler("goals", goals_command))
    app.add_handler(CommandHandler("bill", bill_command))
    app.add_handler(CommandHandler("bills", bills_command))
    app.add_handler(CommandHandler("addcategory", addcategory_command))
    app.add_handler(CommandHandler("categories", categories_command))
    app.add_handler(CommandHandler("split", split_command))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    app.run_polling()

if __name__ == "__main__":
    main()